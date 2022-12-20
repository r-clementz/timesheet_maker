from datetime import datetime
from openpyxl import Workbook,load_workbook 
import getpass

def export_hours(hours,month_total,name,today):
  
    wb=Workbook()
    sheet=wb.create_sheet()
    sheet.title=f"{today.strftime('%B')} {today.strftime('%y')}"
    
    fname=f"{name}'s timesheet.xlsx"

    sheet['A1']="Date"
    sheet['C1']="Start time"
    sheet['D1']="End time"
    sheet['E1']="Duration"
    
    length=len(hours)
    
    total=[]
    for i in range(length):
        sheet[f'A{i+2}']=hours[i]['Date']
        sheet[f'B{i+2}']=hours[i]['Yobi']
        sheet[f'C{i+2}']=hours[i]['Start time']
        sheet[f'D{i+2}']=hours[i]['End time']
        sheet[f'E{i+2}']=hours[i]['Duration']
        
    sheet[f'D{length+2}']="Total"
    sheet[f'E{length+2}']=month_total   
    wb.save(fname)

    print(f"Your hours are exported to Excel file ({name}'s timesheet.xlsx)")    

def edit_data(hours,total_hours,month_total,today,year,month):
    #Reopen excel file to edit 
    fname=f"{name}'s timesheet.xlsx"
    wb_reopen = load_workbook(fname)
    sheet = wb_reopen.get_sheet_by_name(f"{today.strftime('%B')} {today.strftime('%y')}")
    
    for i in range(len(hours)):
        print(f"{i}.{hours[i]['Date']}({hours[i]['Yobi']}) {hours[i]['Start time']}-{hours[i]['End time']}")
    
    index=int(getpass.getpass("\nChoose the date you'd like to edit"))
    
    if index > len(hours):
        index=int(getpass.getpass("Please choose right date again"))
        
    elmnt=getpass.getpass("Which you'd like to edit?\n1. Date 2. Start time 3. End time\n") 
        
    #Edit Date 
    if elmnt=="1":
        old_date=hours[index]['Date']
        new_date=input(f"Old date:{old_date}\nNew date:{year}-{month}-") 
        hours[index]['Date']=f"{old_date[:-2]}{new_date}"

        sheet[f'A{index+2}']=hours[index]['Date']
        
    #Edit Start time     
    elif elmnt=="2":
        new_st=input(f"Start time:{hours[index]['Start time']}\nNew start time:")
        hours[index]['Start time']=new_st
        et=hours[index]['End time']
        hours[index]['Duration']=get_duration(new_st,et)
        total_hours[index]=hours[index]['Duration']
        month_total=sum(total_hours)

        sheet[f'C{index+2}']=hours[index]['Start time']
        sheet[f'E{index+2}']=hours[index]['Duration']
        sheet[f'E{length+1}']=month_total 
        
    #Edit End time 
    elif elmnt=="3":
        new_et=input(f"Start time:{hours[index]['End time']}\nNew end time:")
        hours[index]['End time']=new_et
        st=hours[index]['Start time']
        hours[index]['Duration']=get_duration(st,new_et)
        total_hours[index]=hours[index]['Duration']
        month_total=sum(total_hours)  

        sheet[f'D{index+2}']=hours[index]['End time']
        sheet[f'E{index+2}']=hours[index]['Duration']
        sheet[f'E{length+1}']=month_total   
    
    wb_reopen.save(fname)  

def get_duration(st,et):
    s=datetime.strptime(st,"%H:%M")
    e=datetime.strptime(et,"%H:%M")
    duration=e-s
    duration= (duration.total_seconds()) / (60*60)
    
    return duration         
        







